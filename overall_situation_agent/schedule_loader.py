from __future__ import annotations

import logging
from collections import defaultdict
from datetime import datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def _format_time(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    return text[:5] if len(text) >= 5 and ":" in text else text


def _format_match(home: str, away: str, kick_off: str | None = None) -> str:
    pair = f"{home} vs {away}"
    return f"{kick_off} {pair}" if kick_off else pair


def _matchday_summary(rounds: list[str], matches: list[dict[str, str]]) -> str:
    round_text = "、".join(rounds) if rounds else "比赛日"
    if not matches:
        return round_text
    display_matches = [_format_match(match["home"], match["away"], match.get("time")) for match in matches[:2]]
    if len(matches) == 1:
        return f"{round_text} {display_matches[0]}"
    if len(matches) == 2:
        return f"{round_text} {display_matches[0]}、{display_matches[1]}"
    return f"{round_text} {display_matches[0]}、{display_matches[1]} 等{len(matches)}场"


def _normalize_date(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except TypeError:
            return None
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def load_schedule_context(schedule_input: Path | None) -> dict[str, Any]:
    if schedule_input is None:
        return {
            "status": "missing",
            "source_name": "",
            "days": {},
            "message": "未提供赛程文件，1.2 未标注赛事日。",
        }

    path = schedule_input.resolve()
    if not path.exists():
        return {
            "status": "error",
            "source_name": path.name,
            "days": {},
            "message": f"赛程文件不存在（{path.name}），1.2 未标注赛事日。",
        }

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook["联赛赛程"] if "联赛赛程" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        rows = worksheet.iter_rows(values_only=True)
        title_row = next(rows, None)
        header_row = next(rows, None)
        if not header_row:
            raise ValueError("赛程文件缺少表头。")
        headers = [str(value).strip() if value is not None else "" for value in header_row]
        required_headers = {"轮次", "日期", "主队", "客队", "城市", "时间"}
        if not required_headers.issubset(set(headers)):
            raise ValueError(f"赛程表头不完整，实际表头：{headers}")

        round_idx = headers.index("轮次")
        date_idx = headers.index("日期")
        home_idx = headers.index("主队")
        away_idx = headers.index("客队")
        city_idx = headers.index("城市")
        time_idx = headers.index("时间")

        by_day: dict[str, dict[str, Any]] = defaultdict(lambda: {"rounds": [], "matches": []})
        current_round = ""

        for row in rows:
            if not row or all(value is None or str(value).strip() == "" for value in row):
                continue
            round_value = row[round_idx] if round_idx < len(row) else None
            if round_value is not None and str(round_value).strip():
                current_round = str(round_value).strip()

            match_date = _normalize_date(row[date_idx] if date_idx < len(row) else None)
            home = str(row[home_idx]).strip() if home_idx < len(row) and row[home_idx] else ""
            away = str(row[away_idx]).strip() if away_idx < len(row) and row[away_idx] else ""
            city = str(row[city_idx]).strip() if city_idx < len(row) and row[city_idx] else ""
            kick_off = _format_time(row[time_idx] if time_idx < len(row) else None)
            if not match_date or not home or not away:
                continue

            bucket = by_day[match_date]
            if current_round and current_round not in bucket["rounds"]:
                bucket["rounds"].append(current_round)
            bucket["matches"].append(
                {
                    "home": home,
                    "away": away,
                    "city": city,
                    "time": kick_off,
                }
            )

        schedule_days: dict[str, dict[str, Any]] = {}
        sorted_days = sorted(by_day.keys())
        for day_key, payload in by_day.items():
            rounds = payload["rounds"]
            matches = payload["matches"]
            schedule_days[day_key] = {
                "is_matchday": True,
                "rounds": rounds,
                "matches": matches,
                "match_count": len(matches),
                "match_summary": _matchday_summary(rounds, matches),
            }

        coverage = {
            "start": sorted_days[0] if sorted_days else None,
            "end": sorted_days[-1] if sorted_days else None,
        }
        return {
            "status": "loaded",
            "source_name": path.name,
            "sheet_name": worksheet.title,
            "title": str(title_row[0]).strip() if title_row and title_row[0] else path.stem,
            "days": schedule_days,
            "coverage": coverage,
            "message": (
                f"已加载赛程文件 {path.name}，覆盖 {coverage['start'] or '未知'} 至 "
                f"{coverage['end'] or '未知'}，共 {len(schedule_days)} 个比赛日。"
            ),
        }
    except Exception as exc:
        logger.warning("Failed to load schedule file %s: %s", path, exc)
        return {
            "status": "error",
            "source_name": path.name,
            "days": {},
            "message": f"赛程文件解析失败（{path.name}），1.2 未标注赛事日。",
        }


def enrich_result_with_schedule(result: dict[str, Any], schedule_context: dict[str, Any]) -> dict[str, Any]:
    days = schedule_context.get("days", {})
    for day in result.get("daily", []):
        matchday = days.get(day["date"])
        if matchday:
            day["is_matchday"] = True
            day["matchday"] = matchday
        else:
            day["is_matchday"] = False
            day["matchday"] = None

    result["schedule"] = schedule_context
    return result
