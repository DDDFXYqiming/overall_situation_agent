from __future__ import annotations

import math
import re
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl import load_workbook

from .schema import DATE_FIELDS, MULTI_VALUE_FIELDS, NUMERIC_FIELDS

SPLIT_PATTERN = re.compile(r"[、，,;；|]+")
TIME_PATTERN = re.compile(r"^(\d{1,2}):(\d{1,2})")
FIELD_ALIASES = {
    "工单编号": "gd_identity",
    "省份编码": "province",
    "省份名称": "province_name",
    "服务时间": "service_time",
    "截止时间": "end_time",
    "服务截至时间": "end_time",
    "服务时间到截止时间的耗时（分钟为单位）": "duration_minutes",
    "服务时间到截止时间的耗时(分钟为单位)": "duration_minutes",
    "服务时间到截止时间的耗时（小时为单位）": "duration_hours",
    "服务时间到截止时间的耗时(小时为单位)": "duration_hours",
    "开始时间的月份": "month",
    "月份": "month",
    "日期": "day",
    "时段": "time_period",
    "具体时间（时:分）": "hour",
    "具体时间(时:分)": "hour",
    "工单内容": "content",
    "处理意见（客服回复）": "cs_reply",
    "处理意见(客服回复)": "cs_reply",
    "反馈思路": "feedback_thought",
    "工单投诉内容": "complaint_content",
    "CSP_ID（服务提供商ID）": "csp_id",
    "CSP_ID(服务提供商ID)": "csp_id",
    "CSP_NAME（服务提供商名称）": "csp_name",
    "CSP_NAME(服务提供商名称)": "csp_name",
    "CSP_PROV_ID（服务提供商省份ID）": "csp_prov_id",
    "CSP_PROV_ID(服务提供商省份ID)": "csp_prov_id",
    "CSP_PROV_NAME（服务提供商省份名称）": "csp_prov_name",
    "CSP_PROV_NAME(服务提供商省份名称)": "csp_prov_name",
    "标签组": "label_group",
    "一级标签": "primary_labels",
    "一级标签集合": "primary_labels",
    "二级标签": "secondary_labels",
    "二级标签集合": "secondary_labels",
    "三级标签": "tertiary_labels",
    "三级标签集合": "tertiary_labels",
    "触发场景-赛事/事件": "scene_event",
    "触发场景-情绪": "scene_emotion",
    "触发场景-服务类型": "scene_service_type",
    "触发场景-服务类型（内容判断）": "scene_service_type",
    "触发场景-服务类型(内容判断)": "scene_service_type",
    "洞察维度": "insight_dimension",
    "洞察维度（用得难/烦/亏）": "insight_dimension",
    "洞察维度(用得难/烦/亏)": "insight_dimension",
    "客户关键诉求": "customer_key_appeal",
    "客户关键 诉求": "customer_key_appeal",
    "客户诉求关键词": "customer_keywords",
    "客服关键处理动作": "cs_key_action",
    "客服处理关键词": "cs_keywords",
    "是否有退费诉求": "has_refund_demand",
    "是否有升级投诉倾向": "has_escalation",
    "模型推理说明": "model_reasoning",
    "比赛信息": "match_info",
    "运营举措": "operation_action",
    "隐性需求描述": "latent_need",
    "隐性需求理由": "latent_need_reason",
    "涉及业务/会员类型": "biz_member_cluster",
    "涉及业务/会员类型_聚类": "biz_member_cluster",
    "营销活动页面名称": "marketing_activity_page",
    "营销活动匹配状态": "marketing_activity_match_status",
    "营销活动匹配关键词": "marketing_activity_match_keywords",
    "年龄": "age",
    "性别": "gender",
    "会话ID": "session_id",
    "号码": "phone_number",
    "咪咕产品名称": "product_name",
    "工单处理意见": "cs_reply",
    "工单流水号": "gd_identity",
    "消息内容": "content",
    "消息子类型": "message_subtype",
    "消息类型": "message_type",
    "渠道ID": "channel_id",
    "渠道名称": "channel_name",
    "用户标识": "user_identity",
    "省份": "province_name",
    "结束时间": "end_time",
    "开始时间": "service_time",
    "四个产品层次": "four_product_level",
    "四个层次": "four_product_level",
    "四个运营": "four_operation",
    "四个层次/四个运营": "four_operation",
}
CONTENT_PRIORITY = ("complaint_content", "content", "model_reasoning", "latent_need")
EXPECTED_FIELDS = {
    "gd_identity",
    "province",
    "province_name",
    "service_time",
    "end_time",
    "duration_hours",
    "month",
    "day",
    "time_period",
    "hour",
    "content",
    "complaint_content",
    "cs_reply",
    "primary_labels",
    "secondary_labels",
    "tertiary_labels",
}


def _is_empty(value: Any) -> bool:
    return value is None or (isinstance(value, float) and math.isnan(value)) or str(value).strip() == ""


def _clean_scalar(value: Any) -> Any:
    if _is_empty(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value).strip() if not isinstance(value, (int, float)) else value


def _canonical_field(field: str) -> str:
    return FIELD_ALIASES.get(field.strip(), field.strip())


def _coerce_hour(value: Any) -> int | None:
    cleaned = _clean_scalar(value)
    if cleaned is None:
        return None
    if isinstance(cleaned, (int, float)):
        return int(cleaned)
    match = TIME_PATTERN.match(str(cleaned).strip())
    if not match:
        return _coerce_number(cleaned, "integer")
    return int(match.group(1))


def _split_values(value: Any) -> list[str]:
    cleaned = _clean_scalar(value)
    if cleaned is None:
        return []
    if not isinstance(cleaned, str):
        cleaned = str(cleaned)
    parts = [part.strip() for part in SPLIT_PATTERN.split(cleaned)]
    return [part for part in parts if part and part not in {"无", "不适用", "{}", "[]"}]


def _split_label_group(value: Any) -> list[str]:
    cleaned = _clean_scalar(value)
    if cleaned is None:
        return []
    text = str(cleaned).strip()
    if not text or text in {"无", "不适用", "{}", "[]"}:
        return []
    try:
        payload = json.loads(text)
    except (TypeError, ValueError):
        return _split_values(text)
    if not isinstance(payload, list):
        return _split_values(text)
    groups: list[str] = []
    for item in payload:
        if not isinstance(item, dict):
            continue
        primary = str(item.get("一级标签") or item.get("primary") or "").strip()
        secondary = str(item.get("二级标签") or item.get("secondary") or "").strip()
        tertiary_value = item.get("三级标签") or item.get("tertiary") or []
        if isinstance(tertiary_value, list):
            tertiaries = [str(part).strip() for part in tertiary_value if str(part).strip()]
        else:
            tertiaries = _split_values(tertiary_value)
        if tertiaries:
            groups.extend(" / ".join(part for part in [primary, secondary, tertiary] if part) for tertiary in tertiaries)
        else:
            group = " / ".join(part for part in [primary, secondary] if part)
            if group:
                groups.append(group)
    return groups


def _coerce_number(value: Any, number_type: str) -> int | float | None:
    cleaned = _clean_scalar(value)
    if cleaned is None:
        return None
    try:
        if number_type == "integer":
            return int(float(cleaned))
        return float(cleaned)
    except (TypeError, ValueError):
        return None


def _match_label(match_info: Any) -> str | None:
    cleaned = _clean_scalar(match_info)
    if cleaned is None:
        return None
    text = str(cleaned).strip()
    if not text or text == "{}":
        return None
    try:
        payload = json.loads(text)
    except (TypeError, ValueError):
        return text[:120]
    if not isinstance(payload, dict) or not payload:
        return None
    home = str(payload.get("主队") or payload.get("home") or "").strip()
    away = str(payload.get("客队") or payload.get("away") or "").strip()
    date_value = str(payload.get("日期") or payload.get("date") or "").strip()
    if home and away:
        return f"{date_value} {home} vs {away}".strip()
    return text[:120]


def _resolve_worksheet(workbook, sheet_name: str):
    if sheet_name in workbook.sheetnames:
        return workbook[sheet_name]

    for candidate in workbook.sheetnames:
        worksheet = workbook[candidate]
        first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not first_row:
            continue
        headers = {str(value).strip() for value in first_row if value is not None}
        canonical_headers = {_canonical_field(header) for header in headers}
        if {"gd_identity", "content"}.issubset(canonical_headers) or {"gd_identity", "complaint_content"}.issubset(canonical_headers):
            return worksheet

    return workbook[workbook.sheetnames[0]]


def _row_to_document(headers: list[str], row: tuple[Any, ...], imported_at: str, source_file: str) -> dict[str, Any] | None:
    doc: dict[str, Any] = {}
    raw_values: dict[str, list[Any]] = {}
    for field, value in zip(headers, row):
        canonical = _canonical_field(field)
        if not canonical:
            continue
        raw_values.setdefault(canonical, []).append(value)

    for field, values in raw_values.items():
        value = next((item for item in values if not _is_empty(item)), None)
        if field == "label_group":
            doc[field] = _split_label_group(value)
        elif field in MULTI_VALUE_FIELDS:
            doc[field] = _split_values(value)
        elif field in DATE_FIELDS:
            doc[field] = _clean_scalar(value)
        elif field == "hour":
            doc[field] = _coerce_hour(value)
        elif field in NUMERIC_FIELDS:
            doc[field] = _coerce_number(value, NUMERIC_FIELDS[field])
        else:
            doc[field] = _clean_scalar(value)

    for field in CONTENT_PRIORITY:
        if doc.get(field):
            doc["content"] = doc[field]
            break
    if doc.get("duration_minutes") is not None and doc.get("duration_hours") is None:
        doc["duration_hours"] = round(float(doc["duration_minutes"]) / 60, 4)
    if doc.get("match_info"):
        label = _match_label(doc.get("match_info"))
        if label:
            doc["match_label"] = [label]

    if not doc.get("gd_identity") and not doc.get("content"):
        return None

    doc["source_file"] = source_file
    doc["imported_at"] = imported_at
    return doc


def iter_tagged_feedback(path: Path, sheet_name: str = "打标结果") -> tuple[int, Iterable[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = _resolve_worksheet(workbook, sheet_name)
    row_iter = worksheet.iter_rows(values_only=True)

    try:
        header_row = next(row_iter)
    except StopIteration:
        workbook.close()
        return 0, iter(())

    headers = [str(value).strip() if value is not None else "" for value in header_row]
    second_row = next(row_iter, None)
    has_display_row = bool(second_row) and str(second_row[0] or "").strip() == "工单编号"
    pending_first_row = None if has_display_row else second_row
    total_records = max(worksheet.max_row - (2 if has_display_row else 1), 0)
    imported_at = datetime.now(timezone.utc).isoformat()

    def _generator() -> Iterable[dict[str, Any]]:
        try:
            if pending_first_row is not None:
                doc = _row_to_document(headers, pending_first_row, imported_at, path.name)
                if doc is not None:
                    yield doc
            for row in row_iter:
                doc = _row_to_document(headers, row, imported_at, path.name)
                if doc is not None:
                    yield doc
        finally:
            workbook.close()

    return total_records, _generator()


def load_tagged_feedback(path: Path, sheet_name: str = "打标结果") -> list[dict]:
    total_records, record_iter = iter_tagged_feedback(path, sheet_name=sheet_name)
    records = list(record_iter)
    if total_records and len(records) != total_records:
        return records
    return records
